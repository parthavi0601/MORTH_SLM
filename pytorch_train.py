"""
PyTorch Training Script — Surface Transport SLM
=================================================
Trains both models from PDFs and optional CSVs:
  Task 1: BiLSTM-CRF NER  → ner_bilstm_crf.pt + ner_vocab.pkl
  Task 2: Neural QA        → qa_neural.pt + qa_vocab.pkl + qa_index.pkl

Usage:
  python pytorch_train.py --pdf-dir data/pdfs
  python pytorch_train.py --pdf-dir data/pdfs --csv-file data/accidents.csv
  python pytorch_train.py --pdf-dir data/pdfs --ner-epochs 50 --qa-epochs 30
"""

import os
import sys
import csv
import argparse
import pickle
import re

from pdf_processor import extract_text_from_pdf, clean_text, extract_sentences
from pytorch_ner import (
    Vocabulary, auto_annotate, create_training_data_from_texts,
    train_ner, GAZETTEERS, build_gazetteer_index,
)
from pytorch_qa import QAVocabulary, train_qa


def chunk_text(text: str, chunk_size=500, overlap=100) -> list:
    """Split text into overlapping chunks for QA indexing.

    Larger chunks (500 chars) with overlap give the QA model
    better context per passage and reduce mid-sentence cuts.
    """
    sentences = re.split(r'(?<=[.!?])\s+', text)
    chunks = []
    current = ""
    prev_tail = ""  # last few sentences for overlap

    for sent in sentences:
        if len(current) + len(sent) < chunk_size:
            current += " " + sent
        else:
            if len(current.strip()) > 40:
                chunks.append(current.strip())
                # Keep tail for overlap with next chunk
                tail_sents = re.split(r'(?<=[.!?])\s+', current.strip())
                prev_tail = " ".join(tail_sents[-2:]) if len(tail_sents) >= 2 else ""
            # Start next chunk with overlap from previous
            if prev_tail and overlap > 0:
                current = prev_tail + " " + sent
            else:
                current = sent
    if len(current.strip()) > 40:
        chunks.append(current.strip())
    return chunks


def main():
    parser = argparse.ArgumentParser(description="Train Surface Transport SLM (PyTorch)")
    parser.add_argument("--pdf-dir", type=str, default="data/pdfs",
                        help="Directory containing PDF files")
    parser.add_argument("--csv-file", type=str, default=None,
                        help="CSV file with accident data (or put files in data/ folder)")
    parser.add_argument("--data-dir", type=str, default="data",
                        help="Directory to scan for CSV/XLSX files")
    parser.add_argument("--models-dir", type=str, default="models",
                        help="Directory to save trained models")
    parser.add_argument("--ner-epochs", type=int, default=50,
                        help="Training epochs for NER model")
    parser.add_argument("--qa-epochs", type=int, default=30,
                        help="Training epochs for QA model")
    parser.add_argument("--ner-lr", type=float, default=0.005)
    parser.add_argument("--qa-lr", type=float, default=0.001)
    parser.add_argument("--ner-hidden", type=int, default=128)
    parser.add_argument("--ner-embed", type=int, default=128)
    parser.add_argument("--max-samples", type=int, default=3000,
                        help="Max NER training samples per epoch (default 3000, "
                             "reduce to 1000 for faster training)")
    args = parser.parse_args()

    os.makedirs(args.models_dir, exist_ok=True)

    # ══════════════════════════════════════════════════════════
    # STEP 1: Extract text from PDFs
    # ══════════════════════════════════════════════════════════

    pdf_files = []
    if os.path.isdir(args.pdf_dir):
        pdf_files = [f for f in os.listdir(args.pdf_dir) if f.lower().endswith('.pdf')]

    if not pdf_files:
        print(f"No PDFs found in {args.pdf_dir}")
        print("Please add PDF files (Motor Vehicles Act, MoRTH reports, etc.)")
        sys.exit(1)

    print(f"\n{'='*70}")
    print(f"  SURFACE TRANSPORT SLM — PyTorch TRAINING PIPELINE")
    print(f"  PDFs: {len(pdf_files)} | NER epochs: {args.ner_epochs} | "
          f"QA epochs: {args.qa_epochs}")
    print(f"{'='*70}\n")

    all_sentences = []
    all_chunks = []
    chunk_sources = []
    doc_registry = []

    print("[STEP 1/5] Extracting text from PDFs...")
    for pdf_file in pdf_files:
        pdf_path = os.path.join(args.pdf_dir, pdf_file)
        print(f"  📄 {pdf_file}...", end=" ")

        result = extract_text_from_pdf(pdf_path)
        if "error" in result:
            print(f"SKIP ({result['error']})")
            continue

        cleaned = clean_text(result["full_text"])
        if len(cleaned) < 50:
            print(f"SKIP (too short)")
            continue

        sentences = extract_sentences(cleaned)
        chunks = chunk_text(cleaned)

        all_sentences.extend(sentences)
        all_chunks.extend(chunks)
        chunk_sources.extend([pdf_file] * len(chunks))
        doc_registry.append({
            "filename": pdf_file,
            "num_pages": len(result["pages"]),
            "num_chars": len(cleaned),
            "num_sentences": len(sentences),
            "num_chunks": len(chunks),
        })
        print(f"→ {len(sentences)} sentences, {len(chunks)} chunks")

    total_chars = sum(d['num_chars'] for d in doc_registry)
    print(f"\n  Total: {len(doc_registry)} docs, {len(all_sentences)} sentences, "
          f"{total_chars:,} chars, {len(all_chunks)} QA chunks\n")

    # ── Add CSV/Excel data if provided ──
    csv_sentences = []
    learned_locations = set()
    learned_vehicles = set()

    # Find all CSV and Excel files in data directory
    data_dir = args.data_dir if hasattr(args, 'data_dir') and args.data_dir else "data"
    data_files = []

    # Check explicit csv-file argument
    if args.csv_file and os.path.isfile(args.csv_file):
        data_files.append(args.csv_file)

    # Also scan data/ folder for any CSV/XLSX files
    if os.path.isdir(data_dir):
        for fname in os.listdir(data_dir):
            fpath = os.path.join(data_dir, fname)
            if fname.lower().endswith(('.csv', '.xlsx')) and fpath not in data_files:
                data_files.append(fpath)

    if data_files:
        print(f"[STEP 1b] Loading structured data from {len(data_files)} file(s)...")

        for data_file in data_files:
            print(f"  📊 {os.path.basename(data_file)}...", end=" ")

            if data_file.lower().endswith('.csv'):
                # ── READ CSV (raw news text) ──
                try:
                    with open(data_file, 'r', encoding='utf-8', errors='ignore') as f:
                        reader = csv.DictReader(f)
                        row_count = 0
                        for row in reader:
                            # Raw article text (your 2022/2023 CSVs have "content" column)
                            content = row.get("content", row.get("Content", row.get("\ufeffcontent", "")))
                            if content and len(content) > 30:
                                csv_sentences.append(content)
                                row_count += 1

                            # Also extract structured fields if present
                            loc = row.get("Location", "").strip()
                            state = row.get("State", "").strip()
                            v1 = row.get("Vehicle 1", row.get("Vehicle1", "")).strip()
                            v2 = row.get("Vehicle/Object 2", row.get("Vehicle2", "")).strip()

                            if loc and len(loc) > 2:
                                learned_locations.add(loc)
                            if v1:
                                learned_vehicles.add(v1.lower())
                            if v2:
                                learned_vehicles.add(v2.lower())

                    print(f"→ {row_count} text articles loaded")
                except Exception as e:
                    print(f"ERROR: {e}")

            elif data_file.lower().endswith('.xlsx'):
                # ── READ EXCEL (structured crash data) ──
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(data_file, read_only=True)
                    ws = wb[wb.sheetnames[0]]

                    # Get header row
                    rows_iter = ws.iter_rows(values_only=True)
                    header = [str(h).strip() if h else "" for h in next(rows_iter)]

                    # Find column indices
                    col_map = {}
                    for i, h in enumerate(header):
                        h_clean = h.lower().strip()
                        if 'location' in h_clean:
                            col_map['location'] = i
                        elif 'state' in h_clean:
                            col_map['state'] = i
                        elif 'vehicle 1' in h_clean or 'vehicle1' in h_clean:
                            col_map['vehicle1'] = i
                        elif 'vehicle' in h_clean and ('2' in h_clean or 'object' in h_clean):
                            col_map['vehicle2'] = i
                        elif h_clean == 'killed':
                            col_map['killed'] = i
                        elif h_clean == 'injured':
                            col_map['injured'] = i
                        elif 'road type' in h_clean:
                            col_map['road_type'] = i
                        elif 'crash type' in h_clean:
                            col_map['crash_type'] = i

                    row_count = 0
                    for row in rows_iter:
                        row_count += 1
                        row_list = list(row)

                        # Learn location names
                        if 'location' in col_map:
                            loc = str(row_list[col_map['location']] or "").strip()
                            if loc and len(loc) > 2 and loc != "Nil":
                                learned_locations.add(loc)

                        # Learn vehicle types
                        for vkey in ['vehicle1', 'vehicle2']:
                            if vkey in col_map:
                                v = str(row_list[col_map[vkey]] or "").strip()
                                if v and len(v) > 1 and v != "Nil":
                                    learned_vehicles.add(v.lower())

                        # Build a natural language sentence from structured data
                        parts = []
                        v1 = str(row_list[col_map.get('vehicle1', 0)] or "").strip() if 'vehicle1' in col_map else ""
                        v2 = str(row_list[col_map.get('vehicle2', 0)] or "").strip() if 'vehicle2' in col_map else ""
                        loc = str(row_list[col_map.get('location', 0)] or "").strip() if 'location' in col_map else ""
                        state = str(row_list[col_map.get('state', 0)] or "").strip() if 'state' in col_map else ""
                        killed = str(row_list[col_map.get('killed', 0)] or "0").strip() if 'killed' in col_map else "0"
                        injured = str(row_list[col_map.get('injured', 0)] or "0").strip() if 'injured' in col_map else "0"
                        crash_type = str(row_list[col_map.get('crash_type', 0)] or "").strip() if 'crash_type' in col_map else ""

                        if v1 and v1 != "Nil":
                            if v2 and v2 != "Nil":
                                parts.append(f"A {v1} and a {v2} were involved in a {crash_type} accident" if crash_type else f"A {v1} and a {v2} were involved in an accident")
                            else:
                                parts.append(f"A {v1} was involved in an accident")
                        if loc and loc != "Nil":
                            parts.append(f"near {loc}")
                        if state and state != "Nil":
                            parts.append(f"in {state}")
                        if killed and killed not in ("0", "0.0", "Nil", ""):
                            parts.append(f"killing {int(float(killed))} persons")
                        if injured and injured not in ("0", "0.0", "Nil", ""):
                            parts.append(f"injuring {int(float(injured))} persons")

                        if len(parts) >= 2:
                            csv_sentences.append(" ".join(parts) + ".")

                    wb.close()
                    print(f"→ {row_count} structured rows, "
                          f"{len(learned_locations)} locations, "
                          f"{len(learned_vehicles)} vehicle types")
                except ImportError:
                    print("SKIP (install openpyxl: pip install openpyxl)")
                except Exception as e:
                    print(f"ERROR: {e}")

        print(f"\n  CSV/Excel totals:")
        print(f"    Sentences added:    {len(csv_sentences)}")
        print(f"    Locations learned:  {len(learned_locations)}")
        print(f"    Vehicle types:      {len(learned_vehicles)}")
        all_sentences.extend(csv_sentences)

    # ══════════════════════════════════════════════════════════
    # STEP 2: Create NER training data (auto-annotated)
    # ══════════════════════════════════════════════════════════

    print(f"\n[STEP 2/5] Creating auto-annotated NER training data...")

    # Inject learned locations/vehicles into gazetteers — O(1) set-based dedup
    from pytorch_ner import GAZETTEERS as NER_GAZETTEERS
    if learned_locations:
        existing_locs = {l.lower() for l in NER_GAZETTEERS.get("LOCATION", [])}
        new_locs = [l.lower() for l in learned_locations if l.lower() not in existing_locs]
        NER_GAZETTEERS.setdefault("LOCATION", []).extend(new_locs)
        print(f"  Injected {len(learned_locations)} Excel locations into NER gazetteers")
    if learned_vehicles:
        existing_vehs = {v.lower() for v in NER_GAZETTEERS.get("VEHICLE", [])}
        new_vehs = [v for v in learned_vehicles if v not in existing_vehs]
        NER_GAZETTEERS["VEHICLE"].extend(new_vehs)
        print(f"  Injected {len(learned_vehicles)} Excel vehicle types into NER gazetteers")

    # Build fast trie index ONCE here (workers inherit it via fork on Linux)
    build_gazetteer_index()

    ner_training_data = create_training_data_from_texts(all_sentences)
    print(f"  -> {len(ner_training_data)} annotated sequences "
          f"(from {len(all_sentences)} sentences)")

    # Show annotation stats
    from collections import Counter
    tag_counter = Counter()
    for _, tags in ner_training_data:
        tag_counter.update(t for t in tags if t != "O")
    print(f"  -> Entity tag distribution:")
    for tag, count in tag_counter.most_common(15):
        print(f"       {tag:25s}: {count}")

    # ══════════════════════════════════════════════════════════
    # STEP 3: Build vocabularies
    # ══════════════════════════════════════════════════════════

    print(f"\n[STEP 3/5] Building vocabularies...")

    # NER vocab (from all training tokens)
    ner_vocab = Vocabulary(min_freq=1)
    all_token_lists = [tokens for tokens, _ in ner_training_data]
    # Also add ungrouped sentences
    for sent in all_sentences:
        all_token_lists.append(re.findall(r'\b\w+\b', sent))
    ner_vocab.build(all_token_lists)
    print(f"  NER vocabulary: {ner_vocab.size} words")

    # QA vocab (from all chunks)
    qa_vocab = QAVocabulary(max_vocab=15000, min_freq=1)
    qa_vocab.build(all_chunks + all_sentences)
    print(f"  QA vocabulary: {qa_vocab.size} words")

    # ══════════════════════════════════════════════════════════
    # STEP 4: Train NER model (BiLSTM-CRF)
    # ══════════════════════════════════════════════════════════

    print(f"\n[STEP 4/5] Training NER model...")
    print(f"  Tip: install tqdm for progress bars → pip install tqdm\n")
    train_ner(
        training_data=ner_training_data,
        vocab=ner_vocab,
        model_dir=args.models_dir,
        epochs=args.ner_epochs,
        lr=args.ner_lr,
        hidden_dim=args.ner_hidden,
        embedding_dim=args.ner_embed,
        max_samples=getattr(args, 'max_samples', 3000),
    )

    # ══════════════════════════════════════════════════════════
    # STEP 5: Train QA model (Neural Retriever)
    # ══════════════════════════════════════════════════════════

    print(f"\n[STEP 5/5] Training QA model...")
    train_qa(
        corpus_chunks=all_chunks,
        chunk_sources=chunk_sources,
        vocab=qa_vocab,
        model_dir=args.models_dir,
        epochs=args.qa_epochs,
        lr=args.qa_lr,
    )

    # ══════════════════════════════════════════════════════════
    # FINAL SUMMARY
    # ══════════════════════════════════════════════════════════

    print(f"\n{'='*70}")
    print(f"  ALL TRAINING COMPLETE")
    print(f"{'='*70}")
    print(f"  Output directory: {os.path.abspath(args.models_dir)}/")
    print()
    for f in sorted(os.listdir(args.models_dir)):
        fpath = os.path.join(args.models_dir, f)
        print(f"  {'├──' if f != sorted(os.listdir(args.models_dir))[-1] else '└──'} "
              f"{f:30s} ({os.path.getsize(fpath):>10,} bytes)")
    print()
    print(f"  Documents:       {len(doc_registry)}")
    print(f"  Total chars:     {total_chars:,}")
    print(f"  NER training:    {len(ner_training_data)} sequences")
    print(f"  QA passages:     {len(all_chunks)}")
    print()
    print(f"  Next: Run the Streamlit app:")
    print(f"    streamlit run pytorch_app.py")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()