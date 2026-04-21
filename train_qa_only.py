"""
Train QA Model Only — Surface Transport SLM
=============================================
Trains ONLY the Neural QA model (Task 2), skipping NER entirely.
Use this when your NER weights are already saved in models/.

Usage:
  python train_qa_only.py
  python train_qa_only.py --qa-epochs 50 --qa-lr 0.001
  python train_qa_only.py --pdf-dir data/pdfs --models-dir models
"""

import os
import sys
import csv
import argparse
import re

from pdf_processor import extract_text_from_pdf, clean_text, extract_sentences
from pytorch_qa import QAVocabulary, train_qa


def chunk_text(text: str, chunk_size=300) -> list:
    """Split text into overlapping chunks for QA indexing."""
    sentences = re.split(r'(?<=[.!?])\s+', text)
    chunks = []
    current = ""
    for sent in sentences:
        if len(current) + len(sent) < chunk_size:
            current += " " + sent
        else:
            if len(current.strip()) > 40:
                chunks.append(current.strip())
            current = sent
    if len(current.strip()) > 40:
        chunks.append(current.strip())
    return chunks


def main():
    parser = argparse.ArgumentParser(description="Train QA Model Only (PyTorch)")
    parser.add_argument("--pdf-dir", type=str, default="data/pdfs",
                        help="Directory containing PDF files")
    parser.add_argument("--csv-file", type=str, default=None,
                        help="CSV file with accident data")
    parser.add_argument("--data-dir", type=str, default="data",
                        help="Directory to scan for CSV/XLSX files")
    parser.add_argument("--models-dir", type=str, default="models",
                        help="Directory to save trained models")
    parser.add_argument("--qa-epochs", type=int, default=30,
                        help="Training epochs for QA model")
    parser.add_argument("--qa-lr", type=float, default=0.001)
    args = parser.parse_args()

    os.makedirs(args.models_dir, exist_ok=True)

    # ══════════════════════════════════════════════════════════
    # STEP 1: Extract text from PDFs
    # ══════════════════════════════════════════════════════════

    pdf_files = []
    if os.path.isdir(args.pdf_dir):
        pdf_files = [f for f in os.listdir(args.pdf_dir) if f.lower().endswith('.pdf')]

    if not pdf_files:
        print(f"No PDFs found in {args.pdf_dir}")
        print("Please add PDF files (Motor Vehicles Act, MoRTH reports, etc.)")
        sys.exit(1)

    print(f"\n{'='*70}")
    print(f"  QA-ONLY TRAINING — Surface Transport SLM")
    print(f"  PDFs: {len(pdf_files)} | QA epochs: {args.qa_epochs}")
    print(f"{'='*70}\n")

    all_sentences = []
    all_chunks = []
    chunk_sources = []

    print("[STEP 1/3] Extracting text from PDFs...")
    for pdf_file in pdf_files:
        pdf_path = os.path.join(args.pdf_dir, pdf_file)
        print(f"  📄 {pdf_file}...", end=" ")

        result = extract_text_from_pdf(pdf_path)
        if "error" in result:
            print(f"SKIP ({result['error']})")
            continue

        cleaned = clean_text(result["full_text"])
        if len(cleaned) < 50:
            print(f"SKIP (too short)")
            continue

        sentences = extract_sentences(cleaned)
        chunks = chunk_text(cleaned)

        all_sentences.extend(sentences)
        all_chunks.extend(chunks)
        chunk_sources.extend([pdf_file] * len(chunks))
        print(f"→ {len(sentences)} sentences, {len(chunks)} chunks")

    print(f"\n  Total: {len(all_sentences)} sentences, {len(all_chunks)} QA chunks\n")

    # ── Add CSV/Excel text if available ──
    csv_sentences = []
    data_dir = args.data_dir if args.data_dir else "data"
    data_files = []

    if args.csv_file and os.path.isfile(args.csv_file):
        data_files.append(args.csv_file)

    if os.path.isdir(data_dir):
        for fname in os.listdir(data_dir):
            fpath = os.path.join(data_dir, fname)
            if fname.lower().endswith(('.csv', '.xlsx')) and fpath not in data_files:
                data_files.append(fpath)

    if data_files:
        print(f"[STEP 1b] Loading structured data from {len(data_files)} file(s)...")

        for data_file in data_files:
            print(f"  📊 {os.path.basename(data_file)}...", end=" ")

            if data_file.lower().endswith('.csv'):
                try:
                    with open(data_file, 'r', encoding='utf-8', errors='ignore') as f:
                        reader = csv.DictReader(f)
                        row_count = 0
                        for row in reader:
                            content = row.get("content", row.get("Content", row.get("\ufeffcontent", "")))
                            if content and len(content) > 30:
                                csv_sentences.append(content)
                                row_count += 1
                    print(f"→ {row_count} text articles loaded")
                except Exception as e:
                    print(f"ERROR: {e}")

            elif data_file.lower().endswith('.xlsx'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(data_file, read_only=True)
                    ws = wb[wb.sheetnames[0]]
                    rows_iter = ws.iter_rows(values_only=True)
                    header = [str(h).strip() if h else "" for h in next(rows_iter)]

                    col_map = {}
                    for i, h in enumerate(header):
                        h_clean = h.lower().strip()
                        if 'location' in h_clean:
                            col_map['location'] = i
                        elif 'state' in h_clean:
                            col_map['state'] = i
                        elif 'vehicle 1' in h_clean or 'vehicle1' in h_clean:
                            col_map['vehicle1'] = i
                        elif 'vehicle' in h_clean and ('2' in h_clean or 'object' in h_clean):
                            col_map['vehicle2'] = i
                        elif h_clean == 'killed':
                            col_map['killed'] = i
                        elif h_clean == 'injured':
                            col_map['injured'] = i
                        elif 'crash type' in h_clean:
                            col_map['crash_type'] = i

                    row_count = 0
                    for row in rows_iter:
                        row_count += 1
                        row_list = list(row)

                        parts = []
                        v1 = str(row_list[col_map.get('vehicle1', 0)] or "").strip() if 'vehicle1' in col_map else ""
                        v2 = str(row_list[col_map.get('vehicle2', 0)] or "").strip() if 'vehicle2' in col_map else ""
                        loc = str(row_list[col_map.get('location', 0)] or "").strip() if 'location' in col_map else ""
                        state = str(row_list[col_map.get('state', 0)] or "").strip() if 'state' in col_map else ""
                        killed = str(row_list[col_map.get('killed', 0)] or "0").strip() if 'killed' in col_map else "0"
                        injured = str(row_list[col_map.get('injured', 0)] or "0").strip() if 'injured' in col_map else "0"
                        crash_type = str(row_list[col_map.get('crash_type', 0)] or "").strip() if 'crash_type' in col_map else ""

                        if v1 and v1 != "Nil":
                            if v2 and v2 != "Nil":
                                parts.append(f"A {v1} and a {v2} were involved in a {crash_type} accident" if crash_type else f"A {v1} and a {v2} were involved in an accident")
                            else:
                                parts.append(f"A {v1} was involved in an accident")
                        if loc and loc != "Nil":
                            parts.append(f"near {loc}")
                        if state and state != "Nil":
                            parts.append(f"in {state}")
                        if killed and killed not in ("0", "0.0", "Nil", ""):
                            parts.append(f"killing {int(float(killed))} persons")
                        if injured and injured not in ("0", "0.0", "Nil", ""):
                            parts.append(f"injuring {int(float(injured))} persons")

                        if len(parts) >= 2:
                            csv_sentences.append(" ".join(parts) + ".")

                    wb.close()
                    print(f"→ {row_count} structured rows")
                except ImportError:
                    print("SKIP (install openpyxl: pip install openpyxl)")
                except Exception as e:
                    print(f"ERROR: {e}")

        print(f"\n  CSV/Excel sentences added: {len(csv_sentences)}")
        all_sentences.extend(csv_sentences)

    # ══════════════════════════════════════════════════════════
    # STEP 2: Build QA vocabulary
    # ══════════════════════════════════════════════════════════

    print(f"\n[STEP 2/3] Building QA vocabulary...")
    qa_vocab = QAVocabulary(max_vocab=15000, min_freq=1)
    qa_vocab.build(all_chunks + all_sentences)
    print(f"  QA vocabulary: {qa_vocab.size} words")

    # ══════════════════════════════════════════════════════════
    # STEP 3: Train QA model (Neural Retriever)
    # ══════════════════════════════════════════════════════════

    print(f"\n[STEP 3/3] Training QA model...")
    train_qa(
        corpus_chunks=all_chunks,
        chunk_sources=chunk_sources,
        vocab=qa_vocab,
        model_dir=args.models_dir,
        epochs=args.qa_epochs,
        lr=args.qa_lr,
    )

    # ══════════════════════════════════════════════════════════
    # DONE
    # ══════════════════════════════════════════════════════════

    print(f"\n{'='*70}")
    print(f"  QA TRAINING COMPLETE")
    print(f"{'='*70}")
    print(f"  Output directory: {os.path.abspath(args.models_dir)}/")
    print()
    for f in sorted(os.listdir(args.models_dir)):
        fpath = os.path.join(args.models_dir, f)
        print(f"  {'├──' if f != sorted(os.listdir(args.models_dir))[-1] else '└──'} "
              f"{f:30s} ({os.path.getsize(fpath):>10,} bytes)")
    print()
    print(f"  QA passages: {len(all_chunks)}")
    print(f"\n  Next: Run the Streamlit app:")
    print(f"    streamlit run pytorch_app.py")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
